import numpy as np
import pandas as pd
import logging
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException


# --- Logging Setup ---

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    filename='update_tool.log',
    filemode='a'
)

# --- Centralized Configurations ---

TEMPLATE_CONFIGS = {
    "Service Contract DLT": {
        "column_mapping": {
            'Agreement_ID__c': 'Temp ID',
            'SVMXC__Company__c': 'UCM__Id__c'
        },
        "default_values": {
            "SVMXC__Active__c": "TRUE",
            "HCS_Status__c": "Draft",
            "HCS_Related_To__c": "Service Contract"
        }
    },
    "Covered Product_Service Contract": {
        "column_mapping": {
            "SVMXC__Installed_Product__c": "Asset#"
        },
        "default_values": {
            "PM_Creation_Status__c": "BATCH"
        }
    },
    "Warranty Service Contract": {
        "column_mapping": {
            'Agreement_ID__c': 'Temp ID'
        },
        "default_values": {
            "SVMXC__Active__c": "TRUE",
            "HCS_Status__c": "Draft",
            "HCS_Related_To__c": "Service Contract"
        }
    },
    "Warranty Covered Product_Service Contract": {
        "column_mapping": {
            "SVMXC__Installed_Product__c": "Asset#"
        },
        "default_values": {
            "PM_Creation_Status__c": "BATCH"
        }
    },
    "Service Plan": {
        "default_values": {
            'SVMXC__Active__c': 'TRUE',
            'Business_Unit__c': 'HCS',
            'HCS_Related_To__c': 'Service Contract',
            'GS_Rev_Rec_Method__c': 'Straight Line',
            'Extend_to_End_of_Month__c': 'FALSE',
            'SVMXC__Labor_Rounding_Type__c': 'Actuals',
            'SVMXC__Travel_Rounding_Type__c': 'Actuals'
        }
    },
    "Service Offering": {
        "column_mapping": {
            'SVMXC__Service_Plan__c': 'Temp ID'
        },
        "default_values": {
            'Billing_Type__c': 'Non-Billable',
            'Business_Unit__c': 'HCS',
            'Work_Order_Type__c': 'Field Service; Remote Service; Vendor; Depot Repair; Service Task'
        }
    },
    "Parts Pricing": {
        "column_mapping": {
            'Pricing_Type__c': 'Coverage_Type__c',
            'Service_Offering__c': 'SVMXC__Available_Services__c (Name)',
            'Unit_Type__c': 'GEHCS_Unit_Type__c'
        },
        "default_values": {
            'Display_Entitlement__c': 'TRUE'
        }
    },
    "Labor Pricing": {
        "column_mapping": {
            'Service_Offering__c': 'SVMXC__Available_Services__c (Name)',
            'Unit_Type__c': 'GEHCS_Unit_Type__c'
        },
        "default_values": {}
    }
}

# --- Validation Rules ---

TEMPLATE_VALIDATION_RULES = {
    "Service Contract DLT": {
        "required": ["Temp ID", "SVMXC__End_Date__c", "SVMXC__Start_Date__c"],
        "unique": ["Temp ID"],
        "types": {"Temp ID": str},
        "values": {"SVMXC__Active__c": "TRUE", "HCS_Status__c": "Draft", "HCS_Related_To__c": "Service Contract"}
    },
    "Service Plan": {
        "required": ["Name"],
        "unique": ["Name"],
        "types": {"Name": str},
        "values": {"SVMXC__Active__c": "TRUE"}
    },
    "Service Offering": {
        "required": ["Temp ID"],
        "unique": ["Temp ID"],
        "types": {"Temp ID": str},
        "values": {"Billing_Type__c": "Non-Billable"}
    },
    "Parts Pricing": {
        "required": ["Service_Offering__c", "Pricing_Type__c"],
        "unique": ["Service_Offering__c"],
        "types": {"Service_Offering__c": str, "Pricing_Type__c": str},
        "values": {"Display_Entitlement__c": "TRUE"}
    },
    "Labor Pricing": {
        "required": ["Service_Offering__c"],
        "unique": ["Service_Offering__c"],
        "types": {"Service_Offering__c": str}
    }
}

# --- Utility Functions ---

def convert_to_date_only(df):
    """Convert all datetime columns in a DataFrame to date only."""
    for col in df.columns:
        try:
            temp_series = pd.to_datetime(df[col], errors='coerce')
            if pd.api.types.is_datetime64_any_dtype(temp_series):
                df[col] = temp_series.dt.date
        except Exception as e:
            logging.debug(f"Column '{col}' is not a datetime type, skipping date conversion: {e}")
    return df

def auto_adjust_columns(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        worksheet.column_dimensions[column].width = max_length + 2

def normalize_value(val):
    if pd.isna(val):
        return val
    if isinstance(val, bool):
        return str(val).upper()
    return str(val).strip().upper()

def is_blank(val):
    if pd.isna(val):
        return True
    if isinstance(val, str) and val.strip() == '':
        return True
    return False

def validate_excel_file(file_path):
    if not os.path.exists(file_path):
        logging.error(f"File not found: {file_path}")
        return False, f"File not found: {file_path}"
    try:
        pd.read_excel(file_path, nrows=1)
        return True, "Valid Excel file."
    except Exception as e:
        logging.error(f"Invalid Excel file: {e}")
        return False, f"Invalid Excel file: {e}"

# --- Validate Template Logic - Summary Sheet ---

def validate_template_logic(template_path, sheet_name='INSERT', templates_to_validate=None, source_df=None):
    template_validation_rules = TEMPLATE_VALIDATION_RULES

    if templates_to_validate is None:
        templates_to_validate = list(template_validation_rules.keys())
    filtered_validation_rules = {
        key: template_validation_rules[key]
        for key in templates_to_validate
        if key in template_validation_rules
    }
    all_validation_passed = True
    consolidated_issues = []
    summary_sheet_name = 'Validation_Summary'
    total_records = 0
    try:
        wb = load_workbook(template_path)
    except FileNotFoundError:
        logging.error(f"Template file not found at: {template_path}")
        return {'issues_df': None, 'total_records': total_records, 'duplicate_temp_id_count': 0, 'default_mismatch_count': 0, 'validation_passed': False}
    except InvalidFileException:
        logging.error(f"Invalid Excel file format for '{template_path}'.")
        return {'issues_df': None, 'total_records': total_records, 'duplicate_temp_id_count': 0, 'default_mismatch_count': 0, 'validation_passed': False}
    if summary_sheet_name in wb.sheetnames:
        del wb[summary_sheet_name]
    unique_templates_to_validate = set(templates_to_validate)
    for template_name in unique_templates_to_validate:
        try:
            df = pd.read_excel(template_path, sheet_name=sheet_name, header=1)
            df.dropna(how='all', inplace=True)
            df.reset_index(inplace=True, drop=True)
            if df.empty:
                logging.info(f"Skipping validation for template '{template_name}'. No data found.")
                continue
            total_records += len(df)
            validation_rules = filtered_validation_rules.get(template_name)
            config = TEMPLATE_CONFIGS.get(template_name, {})
            default_values = config.get("default_values", {})
            relevant_columns = [col for col in df.columns if pd.Series(df[col]).notna().any() and col is not None]
            key_columns = [col for col in validation_rules.get('unique', []) if col in relevant_columns]          
            # --- Validate source values against defaults ---
            if source_df is not None and not source_df.empty:
                for col, expected in default_values.items():
                    if col in source_df.columns:
                        for idx, val in source_df[col].items():
                            if pd.notnull(val) and normalize_value(val) != normalize_value(expected):
                                consolidated_issues.append({
                                    'Template Name': template_name,
                                    'Row_Index': idx,
                                    'Temp ID': source_df.loc[idx].get('Temp ID', 'N/A'),
                                    'Column Name': col,
                                    'Issue': f"Source value '{val}' does not match default '{expected}'"
                                })
            # Duplicate check

            if key_columns:
                duplicates_mask = df.duplicated(subset=key_columns, keep=False)
                for index, row in df[duplicates_mask].iterrows():
                    consolidated_issues.append({
                        'Template Name': template_name,
                        'Row_Index': index,
                        'Temp ID': row.get('Temp ID', 'N/A'),
                        'Column Name': key_columns[0],
                        'Issue': 'Duplicate entry Identified'
                    })
            # Required fields
            for col in validation_rules.get("required", []):
                blank_rows = df.index[df[col].apply(is_blank)]
                for blank_index in blank_rows:
                    consolidated_issues.append({
                        'Template Name': template_name,
                        'Row_Index': blank_index,
                        'Temp ID': df.loc[blank_index].get('Temp ID', 'N/A'),
                        'Column Name': col,
                        'Issue': f'{col} must not contain blank'
                    })
            # Type checks
            for col, typ in validation_rules.get("types", {}).items():
                for idx, val in df[col].items():
                    if pd.notnull(val) and not isinstance(val, typ):
                        consolidated_issues.append({
                            'Template Name': template_name,
                            'Row_Index': idx,
                            'Temp ID': df.loc[idx].get('Temp ID', 'N/A'),
                            'Column Name': col,
                            'Issue': f"{col} should be {typ.__name__}"
                        })
            # Value checks (default value mismatches: compare mapped value to default)
            for col, expected in validation_rules.get("values", {}).items():
                for idx, val in df[col].items():
                    if pd.notnull(val) and normalize_value(val) != normalize_value(expected):
                        consolidated_issues.append({
                            'Template Name': template_name,
                            'Row_Index': idx,
                            'Temp ID': df.loc[idx].get('Temp ID', 'N/A'),
                            'Column Name': col,
                            'Issue': f"{col} value '{val}' does not match default '{expected}'"
                        })
            # Temp ID blank check
            if 'Temp ID' in df.columns:
                temp_ids = df['Temp ID']
                mask = temp_ids.isna() | (temp_ids.astype(str).str.strip() == '')
                for idx in df.index[mask]:
                    consolidated_issues.append({
                        'Template Name': template_name,
                        'Row_Index': idx,
                        'Temp ID': df.loc[idx].get('Temp ID', 'N/A'),
                        'Column Name': 'Temp ID',
                        'Issue': 'Temp ID is blank'
                    })
            if consolidated_issues:
                all_validation_passed = False
        except Exception as e:
            logging.error(f"Error during validation for template '{template_name}': {e}")
            all_validation_passed = False

    # --- Build summary sheet with issue columns and summary ---
    default_mismatch_count = 0
    duplicate_temp_id_count = 0

    if consolidated_issues:
        issues_df = pd.DataFrame(consolidated_issues)
        df = pd.read_excel(template_path, sheet_name=sheet_name, header=1)
        df = df.reset_index(drop=True)
        df['Row_Index'] = df.index
        issues_df = issues_df.rename(columns={'Row_Index': 'Row_Index'})
        merged = pd.merge(df, issues_df, on='Row_Index', how='left')
        issue_columns = issues_df['Column Name'].unique()
        for col in issue_columns:
            col_issues = merged.loc[merged['Column Name'] == col].groupby('Row_Index')['Issue'].apply(
                lambda x: '; '.join(x.dropna().unique()))
            df[f'Validation_{col}'] = df.index.map(col_issues).fillna('')
        validation_cols = [f'Validation_{col}' for col in issue_columns]
        df['Validation_Summary'] = df[validation_cols].agg(lambda x: '; '.join([i for i in x if i]), axis=1)
        issue_rows = df[validation_cols + ['Validation_Summary']].apply(lambda x: any(i for i in x), axis=1)
        summary_df = df.loc[issue_rows, ['Row_Index','Temp ID'] + validation_cols + ['Validation_Summary']]
        summary_df = summary_df[summary_df['Temp ID'].notna()]
        with pd.ExcelWriter(template_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            summary_df.to_excel(writer, sheet_name=summary_sheet_name, index=False)
        wb = load_workbook(template_path)
        ws = wb[summary_sheet_name]
        auto_adjust_columns(ws)
        wb.save(template_path)
        # Count duplicates and default mismatches from summary sheet
        duplicate_temp_id_count = summary_df['Validation_Summary'].str.contains('Duplicate entry Identified').sum()
        default_mismatch_count = summary_df['Validation_Summary'].str.contains('does not match default').sum()
        return {
            'issues_df': issues_df,
            'total_records': total_records,
            'duplicate_temp_id_count': duplicate_temp_id_count,
            'default_mismatch_count': default_mismatch_count,
            'validation_passed': all_validation_passed
        }
    else:
        return {
            'issues_df': None,
            'total_records': total_records,
            'duplicate_temp_id_count': 0,
            'default_mismatch_count': 0,
            'validation_passed': True
        }
# --- Generic Update Logic ---

def generic_update_logic(template_path, source_df, column_mapping=None, default_values=None, reference_df=None, sheet_name='INSERT'):
    try:
        template_headers = pd.read_excel(template_path, sheet_name=sheet_name, header=1, nrows=0).columns.tolist()
        output_df = pd.DataFrame(index=range(len(source_df)), columns=template_headers)

        # Map columns from source/reference
        if column_mapping:
            for col in template_headers:
                src_col = column_mapping.get(col)
                if src_col and src_col in source_df.columns:
                    output_df[col] = source_df[src_col]
                elif src_col and reference_df is not None and src_col in reference_df.columns:
                    output_df[col] = reference_df[src_col]
                elif col in source_df.columns:
                    output_df[col] = source_df[col]
                elif reference_df is not None and col in reference_df.columns:
                    output_df[col] = reference_df[col]

        # Override with default values
        if default_values:
            for col in template_headers:
                if col in default_values:
                    output_df[col] = [default_values[col]] * len(output_df)

        # --- Temp ID logic ---
        if 'Temp ID' in output_df.columns:
            if 'Temp ID' in source_df.columns:
                mask = output_df['Temp ID'].isna() | (output_df['Temp ID'].astype(str).str.strip() == '')
                if mask.any():
                    output_df.loc[mask, 'Temp ID'] = range(5001, 5001 + mask.sum())
            else:
                output_df['Temp ID'] = range(5001, 5001 + len(output_df))

        # Write to Excel
        with pd.ExcelWriter(template_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            output_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2, header=False)
        logging.info(f"Updated template '{template_path}' on sheet '{sheet_name}' with {len(source_df)} records.")
        return {'status': True, 'record_count': len(source_df)}
    except Exception as e:
        logging.error(f"Error in generic_update_logic: {e}")
        return {'status': False, 'record_count': 0, 'error': str(e)}

# --- Update Logic Functions ---

def update_template(template_type, template_path, source_df, sheet_name='INSERT', reference_df=None):
    config = TEMPLATE_CONFIGS.get(template_type, {})
    column_mapping = config.get("column_mapping", {})
    default_values = config.get("default_values", {})
    logging.info(f"Running update for template '{template_type}'")
    return generic_update_logic(template_path, source_df, column_mapping, default_values, reference_df, sheet_name=sheet_name)

def update_service_plan(template_path, source_df, sheet_name='INSERT'):
    if 'Name' not in source_df.columns:
        logging.error("Source DataFrame must contain a 'Name' column for service plan uniqueness.")
        raise ValueError("Source DataFrame must contain a 'Name' column for service plan uniqueness.")
    unique_df = source_df.drop_duplicates(subset=['Name']).reset_index(drop=True)
    config = TEMPLATE_CONFIGS["Service Plan"]
    default_values = config.get("default_values", {})
    template_headers = pd.read_excel(template_path, sheet_name=sheet_name, header=1, nrows=0).columns.tolist()
    output_df = pd.DataFrame(index=range(len(unique_df)), columns=template_headers)
    for idx, row in unique_df.iterrows():
        name_value = str(row['Name']) if not pd.isna(row['Name']) else ''
        row_defaults = default_values.copy()
        if 'warranty' in name_value.lower():
            row_defaults.update({
                'GS_Rev_Rec_Method__c': 'Warranty',
                'HCS_Related_To__c': 'Warranty',
                'Account_Type__c': 'Customer',
                'Duration_months__c': '12',
                'Start_Date__c': 'eOM Warranty Start Date'
            })
        for col in template_headers:
            if col in row_defaults:
                output_df.at[idx, col] = row_defaults[col]
            elif col in unique_df.columns:
                output_df.at[idx, col] = row[col]
    # --- Temp ID logic for Service Plan ---
    if 'Temp ID' in output_df.columns:
        if 'Temp ID' not in unique_df.columns:
            output_df['Temp ID'] = range(5001, 5001 + len(output_df))
        else:
            mask = output_df['Temp ID'].isna() | (output_df['Temp ID'].astype(str).str.strip() == '')
            if mask.any():
                output_df.loc[mask, 'Temp ID'] = range(5001, 5001 + mask.sum())
    with pd.ExcelWriter(template_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        output_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2, header=False)
    logging.info(f"Updated Service Plan template '{template_path}' with {len(unique_df)} records.")
    return {'status': True, 'record_count': len(unique_df)}

def update_service_offering(template_path, source_df, reference_df=None, sheet_name='INSERT'):
    config = TEMPLATE_CONFIGS["Service Offering"]
    column_mapping = config.get("column_mapping", {})
    default_values = config.get("default_values", {})
    template_headers = pd.read_excel(template_path, sheet_name=sheet_name, header=1, nrows=0).columns.tolist()
    output_df = pd.DataFrame(index=range(len(source_df)), columns=template_headers)
    for idx, row in source_df.iterrows():
        for col in template_headers:
            if col in column_mapping and column_mapping[col] in source_df.columns:
                output_df.at[idx, col] = row[column_mapping[col]]
            elif col in source_df.columns:
                output_df.at[idx, col] = row[col]
            elif col in default_values:
                output_df.at[idx, col] = default_values[col]
            elif reference_df is not None and col in reference_df.columns:
                output_df.at[idx, col] = reference_df.at[idx, col] if idx < len(reference_df) else None
    # --- Temp ID logic for Service Offering ---
    if 'Temp ID' in output_df.columns:
        if 'Temp ID' not in source_df.columns:
            output_df['Temp ID'] = range(5001, 5001 + len(output_df))
        else:
            mask = output_df['Temp ID'].isna() | (output_df['Temp ID'].astype(str).str.strip() == '')
            if mask.any():
                output_df.loc[mask, 'Temp ID'] = range(5001, 5001 + mask.sum())
    with pd.ExcelWriter(template_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        output_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2, header=False)
    logging.info(f"Updated Service Offering template '{template_path}' with {len(source_df)} records.")
    return {'status': True, 'record_count': len(source_df)}

def update_parts_pricing(template_path, source_df, sheet_name='INSERT'):
    config = TEMPLATE_CONFIGS["Parts Pricing"]
    column_mapping = config.get("column_mapping", {})
    default_values = config.get("default_values", {})
    filtered = source_df[source_df.get('Need parts pricing', '').str.lower() == 'yes'].copy()
    if filtered.empty:
        logging.warning("No records found for parts pricing update.")
        return False
    template_headers = pd.read_excel(template_path, sheet_name=sheet_name, header=1, nrows=0).columns.tolist()
    output_df = pd.DataFrame(index=range(len(filtered)), columns=template_headers)
    for idx, row in filtered.iterrows():
        for col in template_headers:
            if col in filtered.columns:
                output_df.at[idx, col] = row[col]
            elif col in column_mapping and column_mapping[col] in filtered.columns:
                output_df.at[idx, col] = row[column_mapping[col]]
            elif col in default_values:
                output_df.at[idx, col] = default_values[col]
    # --- Temp ID logic for Parts Pricing ---
    if 'Temp ID' in output_df.columns:
        if 'Temp ID' not in filtered.columns:
            output_df['Temp ID'] = range(5001, 5001 + len(output_df))
        else:
            mask = output_df['Temp ID'].isna() | (output_df['Temp ID'].astype(str).str.strip() == '')
            if mask.any():
                output_df.loc[mask, 'Temp ID'] = range(5001, 5001 + mask.sum())
    with pd.ExcelWriter(template_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        output_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2, header=False)
    logging.info(f"Updated Parts Pricing template '{template_path}' with {len(filtered)} records.")
    return True

def update_labor_pricing(template_path, source_df, sheet_name='INSERT'):
    config = TEMPLATE_CONFIGS["Labor Pricing"]
    column_mapping = config.get("column_mapping", {})
    default_values = config.get("default_values", {})
    filtered = source_df[source_df.get('Need labor pricing', '').str.lower() == 'yes'].copy()
    if filtered.empty:
        logging.warning("No records found for labor pricing update.")
        return False
    template_headers = pd.read_excel(template_path, sheet_name=sheet_name, header=1, nrows=0).columns.tolist()
    output_df = pd.DataFrame(index=range(len(filtered)*2), columns=template_headers)
    row_idx = 0
    for _, row in filtered.iterrows():
        for labor_type in ['Labor', 'Travel']:
            for col in template_headers:
                if col == 'Labor_Type__c':
                    output_df.at[row_idx, col] = labor_type
                elif col in filtered.columns:
                    output_df.at[row_idx, col] = row[col]
                elif col in column_mapping and column_mapping[col] in filtered.columns:
                    output_df.at[row_idx, col] = row[column_mapping[col]]
                elif col in default_values:
                    output_df.at[row_idx, col] = default_values[col]
            row_idx += 1
    # --- Temp ID logic for Labor Pricing ---
    if 'Temp ID' in output_df.columns:
        if 'Temp ID' not in filtered.columns:
            output_df['Temp ID'] = range(5001, 5001 + len(output_df))
        else:
            mask = output_df['Temp ID'].isna() | (output_df['Temp ID'].astype(str).str.strip() == '')
            if mask.any():
                output_df.loc[mask, 'Temp ID'] = range(5001, 5001 + mask.sum())
    with pd.ExcelWriter(template_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        output_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2, header=False)
    logging.info(f"Updated Labor Pricing template '{template_path}' with {len(filtered)} records.")
    return True

def update_template_with_picklist(template_path, sheet_name, picklist_values):
    """
    Overwrites columns in the template with the selected picklist values.
    picklist_values: dict {column_name: value}
    """
    try:
        wb = load_workbook(template_path)
        ws = wb[sheet_name]
        header_row = [cell.value for cell in ws[2]]
        for col, value in picklist_values.items():
            if value and col in header_row:
                col_idx = header_row.index(col) + 1
                for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.value = value
        wb.save(template_path)
        logging.info(f"Applied picklist values to template '{template_path}' on sheet '{sheet_name}'.")
        return True
    except Exception as e:
        logging.error(f"Error updating picklist values: {e}")
        return False

# --- Mapping Functions ---

def account_location_mapping(source_file_path, mapping_file_path, output_file_path=None):
    """Map account/location data and write results back to Excel."""
    if output_file_path is None:
        output_file_path = source_file_path
    try:
        mapping_df = pd.read_excel(mapping_file_path)
        source_df = pd.read_excel(source_file_path)
        # Example mapping logic (customize as needed)
        source_df['Ship_to_validate'] = np.where(
            source_df['Ship_to_check'].astype(str).str.lower().isin(['yes', 'primary']),
            'Available',
            'Not Available'
        )
        source_df['Bill_to_validate'] = np.where(
            source_df['Bill_to_check'].astype(str).str.lower().isin(['yes', 'primary']),
            'Available',
            'Not Available'
        )
        with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            source_df.to_excel(writer, index=False)
        logging.info(f"Account/location mapping updated: {output_file_path}")
        return True
    except Exception as e:
        logging.error(f"Account/location mapping error: {e}")
        return False

def install_product_mapping(source_file_path, mapping_file_path, output_file_path=None):
    """Map install product data and write results back to Excel."""
    if output_file_path is None:
        output_file_path = source_file_path
    try:
        mapping_df = pd.read_excel(mapping_file_path)
        source_df = pd.read_excel(source_file_path)
        valid_products = set(mapping_df['SVMXC__SM_External_ID__c'].dropna())
        source_df['Install_Product_Status'] = np.where(
            source_df['Asset#'].isin(valid_products),
            'Available',
            'Not Available'
        )
        with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            source_df.to_excel(writer, index=False)
        logging.info(f"Install product mapping updated: {output_file_path}")
        return True
    except Exception as e:
        logging.error(f"Install product mapping error: {e}")
        return False

# --- Validation Functions ---

def location_mapping_validate_fixed(source_path, mapping_file_path):
    """Advanced validation for account/location mapping."""
    summary = {
        'status': False,
        'message': '',
        'source_records': 0,
        'ucm_id_found': 0,
        'ucm_id_not_found': 0,
        'bill_to_valid_count': 0,
        'bill_to_not_valid_count': 0,
        'ship_to_valid_count': 0,
        'ship_to_not_valid_count': 0,
        'updated_records': 0
    }
    try:
        mapping_df = pd.read_excel(mapping_file_path)
        source_df = pd.read_excel(source_path)
        summary['source_records'] = len(source_df)
        # Example validation logic (customize as needed)
        summary['bill_to_valid_count'] = source_df['Bill_to_check'].astype(str).str.lower().isin(['yes', 'primary']).sum()
        summary['bill_to_not_valid_count'] = (source_df['Bill_to_check'].astype(str).str.lower() == 'no').sum()
        summary['ship_to_valid_count'] = source_df['Ship_to_check'].astype(str).str.lower().isin(['yes', 'primary']).sum()
        summary['ship_to_not_valid_count'] = (source_df['Ship_to_check'].astype(str).str.lower() == 'no').sum()
        summary['status'] = True
        summary['message'] = "Validation completed."
        logging.info(f"Location mapping validation summary: {summary}")
        return summary
    except Exception as e:
        summary['message'] = f"Validation error: {e}"
        logging.error(summary['message'])
        return summary

def validate_install_product_mapping(source_path, mapping_file_path):
    """Advanced validation for install product mapping."""
    summary = {
        'status': False,
        'message': '',
        'source_records': 0,
        'updated_records': 0,
        'matched_products': 0,
        'unmatched_products': 0,
    }
    try:
        mapping_df = pd.read_excel(mapping_file_path)
        source_df = pd.read_excel(source_path)
        valid_products = set(mapping_df['SVMXC__SM_External_ID__c'].dropna())
        summary['source_records'] = len(source_df)
        matches = source_df[source_df['Asset#'].isin(valid_products)]
        summary['matched_products'] = len(matches)
        summary['unmatched_products'] = summary['source_records'] - summary['matched_products']
        summary['updated_records'] = summary['matched_products']
        summary['status'] = True
        summary['message'] = "Install product mapping validation completed."
        logging.info(f"Install product mapping validation summary: {summary}")
        return summary
    except Exception as e:
        summary['message'] = f"Validation error: {e}"
        logging.error(summary['message'])
        return summary

# --- Picklist Update ---

def update_template_with_picklist(template_path, sheet_name, picklist_values):
    """Overwrite columns in the template with selected picklist values."""
    try:
        wb = load_workbook(template_path)
        ws = wb[sheet_name]
        header_row = [cell.value for cell in ws[2]]
        for col, value in picklist_values.items():
            if value and col in header_row:
                col_idx = header_row.index(col) + 1
                for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.value = value
        wb.save(template_path)
        logging.info(f"Picklist values applied to template: {template_path}, sheet: {sheet_name}")
        return True
    except Exception as e:
        logging.error(f"Picklist update error: {e}")
        return False
    
# --- End of script ---



























